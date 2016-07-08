import csv
from io import BytesIO
import zipfile

from chardet.universaldetector import UniversalDetector
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from rest_framework.serializers import HyperlinkedModelSerializer
from drf_enum_field.serializers import EnumFieldSerializerMixin
from drf_hal_json.serializers import HalModelSerializer, HalEmbeddedSerializer
from drf_nested_routing.serializers import NestedRoutingSerializerMixin


class HalNestedRoutingEmbeddedSerializer(NestedRoutingSerializerMixin, EnumFieldSerializerMixin, HalEmbeddedSerializer):
    pass


class HalNestedRoutingLinksSerializer(NestedRoutingSerializerMixin, EnumFieldSerializerMixin, HyperlinkedModelSerializer):
    pass


class HalNestedFieldsModelSerializer(NestedRoutingSerializerMixin, EnumFieldSerializerMixin, HalModelSerializer):
    links_serializer_class = HalNestedRoutingLinksSerializer
    embedded_serializer_class = HalNestedRoutingEmbeddedSerializer


class CsvSerializer(object):
    @staticmethod
    def serialize(data, separator='\t'):
        if isinstance(data, bytes):
            return data

        if not isinstance(data, list):
            data = [str(data)]

        csv_buffer = BytesIO()
        for row in data:
            if not isinstance(row, (list, tuple)):
                row = [row]
            csv_buffer.write((separator.join(CsvSerializer.__validate_cell(cell) for cell in row) + '\n').encode('utf-8'))

        return csv_buffer.getvalue()

    @staticmethod
    def deserialize(file_bytes):
        try:
            file_string = file_bytes.decode('utf-8')
        except UnicodeDecodeError as ude:
            detector = UniversalDetector()
            for line in BytesIO(file_bytes):
                detector.feed(line)
                if detector.done:
                    break
            detector.close()
            if detector.result['confidence'] < 0.5:
                raise ValueError("Failed to guess the encoding of the file (it's not utf-8). Use utf-8 encoded files.")
            try:
                file_string = file_bytes.decode(detector.result['encoding'])
            except UnicodeDecodeError:
                raise ValueError("Failed to guess the encoding of the file (it's not utf-8). Use utf-8 encoded files. "
                                 "(The invalid character is '{char:#x}' at {pos})".format(pos=ude.start,
                                                                                          char=file_bytes[ude.start]))
        csv_lines = file_string.splitlines()
        first_line = csv_lines[:1]
        first_row_tab = next(csv.reader(first_line, delimiter="\t"))
        first_row_semicolon = next(csv.reader(first_line, delimiter=";"))
        if len(first_row_tab) > 1:
            rows = csv.reader(csv_lines, delimiter="\t")
        elif len(first_row_semicolon) > 1:
            rows = csv.reader(csv_lines, delimiter=";")
        else:
            raise ValueError("Csv file is not delimited by ';' or 'tab'")

        return rows

    @staticmethod
    def __validate_cell(cell):
        cell = str(cell) if cell is not None else ''
        if "\t" in cell or "\n" in cell or '"' in cell:
            cell = cell.replace('"', '""')
            cell = '"{}"'.format(cell)
        return cell


class ZipSerializer(object):
    @staticmethod
    def serialize(data):
        byte_buffer = BytesIO()
        zip_file = zipfile.ZipFile(byte_buffer, "w")
        for filename, data_bytes in data.items():
            if not isinstance(data_bytes, bytes):
                return data
            zip_file.writestr(filename, data_bytes)
        zip_file.close()
        return byte_buffer.getvalue()


class XlsxSerializer(object):
    @staticmethod
    def serialize(data):
        workbook = Workbook()
        sheet = workbook.active
        for row_index, row in enumerate(data):
            for column_index, value in enumerate(row):
                data_type = Cell.TYPE_STRING
                if isinstance(value, (int, float)):
                    data_type = Cell.TYPE_NUMERIC
                if data_type == Cell.TYPE_STRING:
                    value = str(value)
                sheet.cell(column=column_index + 1, row=row_index + 1).set_explicit_value(value, data_type=data_type)
        xlsx_file = BytesIO()
        workbook.save(xlsx_file)
        return xlsx_file.getvalue()

    @staticmethod
    def deserialize(file_bytes, sheet_name):
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        if len(workbook.worksheets) > 1:
            if not sheet_name:
                raise ValueError("The uploaded file contains several sheets. The name of the sheet to be imported "
                                 "needs to be specified with the 'sheetName' parameter.")
            worksheet = workbook.get_sheet_by_name(sheet_name)
        else:
            worksheet = workbook.get_active_sheet()

        if not worksheet:
            raise ValueError("No worksheet found.")

        return [[cell.value for cell in row] for row in worksheet.rows]
